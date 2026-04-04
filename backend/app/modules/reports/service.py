import calendar
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.fuel.models import FuelDelivery, FuelRefill
from app.modules.generators.models import Generator
from app.modules.motohours.models import MaintenanceLog
from app.modules.shifts.models import Shift
from app.modules.users.models import User

KYIV = ZoneInfo("Europe/Kyiv")

UA_DAYS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд"]
UA_MONTHS = ["", "Січень", "Лютий", "Березень", "Квітень", "Травень", "Червень",
             "Липень", "Серпень", "Вересень", "Жовтень", "Листопад", "Грудень"]

# ── Стилі ──────────────────────────────────────────────────────────────────
HDR1 = Font(bold=True, size=13, name="Calibri")
HDR2 = Font(bold=True, size=11, name="Calibri")
BODY = Font(size=10, name="Calibri")
BOLD = Font(bold=True, size=10, name="Calibri")

FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FILL_SUBHDR  = PatternFill("solid", fgColor="2E75B6")
FILL_ALT     = PatternFill("solid", fgColor="EBF3FB")
FILL_TOTAL   = PatternFill("solid", fgColor="D6E4F0")
FILL_WEEKEND = PatternFill("solid", fgColor="FFF2CC")
FILL_KPI_LBL = PatternFill("solid", fgColor="F2F2F2")

WHITE = Font(bold=True, size=11, name="Calibri", color="FFFFFF")

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _cell(ws, row, col, value=None, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:          c.font = font
    if fill:          c.fill = fill
    if align:         c.alignment = align
    if border:        c.border = border
    if number_format: c.number_format = number_format
    return c


def _merge(ws, r1, c1, r2, c2, value=None, font=None, fill=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if font:  c.font = font
    if fill:  c.fill = fill
    if align: c.alignment = align
    return c


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    # ── HEAD DATA FETCHING ──────────────────────────────────────────────────

    async def _get_generator(self, generator_id: uuid.UUID) -> Generator | None:
        r = await self.db.execute(select(Generator).where(Generator.id == generator_id))
        return r.scalar_one_or_none()

    async def _get_shifts_for_month(self, generator_id: uuid.UUID, year: int, month: int) -> list[Shift]:
        tz_start = datetime(year, month, 1, 0, 0, 0, tzinfo=KYIV)
        last_day = calendar.monthrange(year, month)[1]
        tz_end   = datetime(year, month, last_day, 23, 59, 59, tzinfo=KYIV)
        r = await self.db.execute(
            select(Shift).where(
                and_(
                    Shift.generator_id == generator_id,
                    Shift.started_at >= tz_start,
                    Shift.started_at <= tz_end,
                )
            ).order_by(Shift.started_at)
        )
        return list(r.scalars().all())

    async def _get_refills_for_month(self, generator_id: uuid.UUID, year: int, month: int) -> list[FuelRefill]:
        tz_start = datetime(year, month, 1, 0, 0, 0, tzinfo=KYIV)
        last_day = calendar.monthrange(year, month)[1]
        tz_end   = datetime(year, month, last_day, 23, 59, 59, tzinfo=KYIV)
        r = await self.db.execute(
            select(FuelRefill).where(
                and_(
                    FuelRefill.generator_id == generator_id,
                    FuelRefill.refilled_at >= tz_start,
                    FuelRefill.refilled_at <= tz_end,
                )
            ).order_by(FuelRefill.refilled_at)
        )
        return list(r.scalars().all())

    async def _get_deliveries_for_month(self, year: int, month: int) -> list[FuelDelivery]:
        tz_start = datetime(year, month, 1, 0, 0, 0, tzinfo=KYIV)
        last_day = calendar.monthrange(year, month)[1]
        tz_end   = datetime(year, month, last_day, 23, 59, 59, tzinfo=KYIV)
        r = await self.db.execute(
            select(FuelDelivery).where(
                and_(
                    FuelDelivery.delivered_at >= tz_start,
                    FuelDelivery.delivered_at <= tz_end,
                )
            ).order_by(FuelDelivery.delivered_at)
        )
        return list(r.scalars().all())

    async def _get_maintenance_for_generator(self, generator_id: uuid.UUID) -> list[tuple]:
        r = await self.db.execute(
            select(MaintenanceLog, User.full_name)
            .outerjoin(User, MaintenanceLog.performed_by == User.id)
            .where(MaintenanceLog.generator_id == generator_id)
            .order_by(MaintenanceLog.performed_at.desc())
        )
        return list(r.all())

    async def _get_user_name(self, user_id: uuid.UUID | None) -> str:
        if not user_id:
            return "—"
        r = await self.db.execute(select(User.full_name).where(User.id == user_id))
        name = r.scalar_one_or_none()
        return name or "—"

    # ── MAIN ENTRY POINT ────────────────────────────────────────────────────

    async def generate_monthly_excel(
        self,
        generator_id: uuid.UUID,
        year: int,
        month: int,
        fuel_price: float = 50.0,
    ) -> BytesIO:
        generator  = await self._get_generator(generator_id)
        gen_name   = generator.name if generator else str(generator_id)
        shifts     = await self._get_shifts_for_month(generator_id, year, month)
        refills    = await self._get_refills_for_month(generator_id, year, month)
        deliveries = await self._get_deliveries_for_month(year, month)
        maintenance = await self._get_maintenance_for_generator(generator_id)

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        await self._build_operational_journal(wb, gen_name, year, month, shifts, refills, fuel_price)
        await self._build_monthly_summary(wb, gen_name, year, month, shifts, refills, deliveries, fuel_price)
        await self._build_maintenance_sheet(wb, gen_name, maintenance)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── АРКУШ 1: ОПЕРАЦІЙНИЙ ЖУРНАЛ ─────────────────────────────────────────

    async def _build_operational_journal(
        self, wb: Workbook, gen_name: str, year: int, month: int,
        shifts: list[Shift], refills: list[FuelRefill], fuel_price: float
    ) -> None:
        ws = wb.create_sheet("Операційний журнал")
        month_str  = f"{UA_MONTHS[month]} {year}"
        now_str    = datetime.now(tz=KYIV).strftime("%d.%m.%Y %H:%M")
        days_count = calendar.monthrange(year, month)[1]

        # ── Шапка ──
        ws.merge_cells("A1:R1")
        ws["A1"] = f"Операційний журнал роботи генератора «{gen_name}»  —  {month_str}"
        ws["A1"].font = HDR1
        ws["A1"].alignment = CENTER

        ws.merge_cells("A2:R2")
        ws["A2"] = f"Генератор: «{gen_name}»  |  Місяць: {month_str}  |  Сформовано: {now_str}  |  Рядків: {days_count}"
        ws["A2"].font = BODY
        ws["A2"].alignment = CENTER

        # ── Групові заголовки (рядок 3) ──
        groups = [
            (1,  1,  "#",        FILL_HEADER, WHITE),
            (2,  2,  "",         FILL_HEADER, WHITE),
            (3,  3,  "День",     FILL_HEADER, WHITE),
            (4,  5,  "РАНОК",    FILL_SUBHDR, WHITE),
            (6,  7,  "ДЕНЬ",     FILL_SUBHDR, WHITE),
            (8,  9,  "ВЕЧІР",    FILL_SUBHDR, WHITE),
            (10, 11, "ВИРОБІТОК", FILL_HEADER, WHITE),
            (12, 15, "ПАЛИВО",   FILL_SUBHDR, WHITE),
            (16, 16, "л/год",    FILL_HEADER, WHITE),
            (17, 18, "ДОКУМЕНТИ", FILL_SUBHDR, WHITE),
        ]
        for c1, c2, label, fill, fnt in groups:
            if c1 == c2:
                _cell(ws, 3, c1, label, font=fnt, fill=fill, align=CENTER, border=BORDER)
            else:
                _merge(ws, 3, c1, 3, c2, label, font=fnt, fill=fill, align=CENTER)
                for cc in range(c1, c2 + 1):
                    ws.cell(3, cc).border = BORDER

        # ── Підзаголовки (рядок 4) ──
        sub_cols = [
            (1,  "#"),
            (2,  "Дата"),
            (3,  "День"),
            (4,  "Р↑ Поч."),
            (5,  "Р↓ Кін."),
            (6,  "Д↑ Поч."),
            (7,  "Д↓ Кін."),
            (8,  "В↑ Поч."),
            (9,  "В↓ Кін."),
            (10, "Год."),
            (11, "Оператор"),
            (12, "Пал.⬆, л"),
            (13, "Пал.⬇, л"),
            (14, "Витрата, л"),
            (15, "Заправка, л"),
            (16, "л/год"),
            (17, "Чек #"),
            (18, "Примітки"),
        ]
        for col, label in sub_cols:
            _cell(ws, 4, col, label, font=WHITE, fill=FILL_SUBHDR, align=CENTER, border=BORDER)

        # ── Індекс заправок по датах ──
        refill_by_date: dict[date, list[FuelRefill]] = {}
        for rf in refills:
            d = rf.refilled_at.astimezone(KYIV).date()
            refill_by_date.setdefault(d, []).append(rf)

        # ── Будуємо індекс змін по даті (може бути кілька на день) ──
        shift_by_date: dict[date, list[Shift]] = {}
        for sh in shifts:
            d = sh.started_at.astimezone(KYIV).date()
            shift_by_date.setdefault(d, []).append(sh)

        # ── Дані ──
        total_hours   = Decimal("0")
        total_fuel_in  = Decimal("0")
        total_fuel_out = Decimal("0")
        total_consumed = Decimal("0")
        total_refill   = Decimal("0")

        for day_num in range(1, days_count + 1):
            d    = date(year, month, day_num)
            row  = 4 + day_num
            wd   = d.weekday()  # 0=Mon … 6=Sun
            is_weekend = wd >= 5
            fill = FILL_WEEKEND if is_weekend else (FILL_ALT if day_num % 2 == 0 else None)

            day_shifts  = shift_by_date.get(d, [])
            day_refills = refill_by_date.get(d, [])

            # Виробіток за день
            day_hours    = sum((sh.motohours_accumulated or Decimal("0")) for sh in day_shifts)
            day_consumed = sum((sh.fuel_consumed_liters  or Decimal("0")) for sh in day_shifts)
            day_refill   = sum((rf.liters for rf in day_refills), Decimal("0"))

            # Заправки
            refill_str = f"{float(day_refill):.1f}" if day_refill else "—"
            check_str  = ", ".join(rf_d.check_number or "" for rf_d in day_refills if rf_d.check_number) or "—"

            # Паливо початок/кінець (з першої зміни дня)
            fuel_start = day_shifts[0].fuel_consumed_liters and (
                # fuel_tank_before was not stored on shift — use stock levels from refills or None
                None
            )

            operator = "—"
            if day_shifts:
                op_id = day_shifts[0].started_by
                if op_id:
                    operator = await self._get_user_name(op_id)

            # Ранок/день/вечір — shift start/end times (grouped by 3 shift periods)
            morning_start = morning_end = day_start = day_end = evening_start = evening_end = "—"
            for idx, sh in enumerate(day_shifts[:3]):
                s = sh.started_at.astimezone(KYIV).strftime("%H:%M")
                e = sh.stopped_at.astimezone(KYIV).strftime("%H:%M") if sh.stopped_at else "—"
                if idx == 0: morning_start, morning_end = s, e
                elif idx == 1: day_start, day_end = s, e
                elif idx == 2: evening_start, evening_end = s, e

            hours_str    = f"{float(day_hours):.2f}" if day_hours else ""
            consumed_str = f"{float(day_consumed):.3f}" if day_consumed else ""
            lph_str      = f"{float(day_consumed / day_hours):.3f}" if day_hours and day_hours > 0 else "—"

            # Паливо⬆ (на початок дня)
            # Use last known refill stock_after as proxy
            fuel_up   = "—"
            fuel_down = "—"

            row_data = [
                day_num, d.strftime("%d.%m.%Y"), UA_DAYS[wd],
                morning_start, morning_end,
                day_start,     day_end,
                evening_start, evening_end,
                hours_str, operator,
                fuel_up, fuel_down,
                consumed_str, refill_str,
                lph_str, check_str, "—",
            ]
            for col, val in enumerate(row_data, start=1):
                _cell(ws, row, col, val,
                      font=BOLD if col == 1 else BODY,
                      fill=fill,
                      align=CENTER,
                      border=BORDER)

            total_hours    += day_hours
            total_consumed += day_consumed
            total_refill   += day_refill

        # ── Підсумок ──
        total_row = 4 + days_count + 1
        _merge(ws, total_row, 1, total_row, 9, "ПІДСУМОК МІСЯЦЯ",
               font=BOLD, fill=FILL_TOTAL, align=CENTER)
        for cc in range(1, 10):
            ws.cell(total_row, cc).border = BORDER
        _cell(ws, total_row, 10, f"{float(total_hours):.2f} год", font=BOLD, fill=FILL_TOTAL, align=CENTER, border=BORDER)
        _cell(ws, total_row, 11, "", fill=FILL_TOTAL, border=BORDER)
        _cell(ws, total_row, 12, "", fill=FILL_TOTAL, border=BORDER)
        _cell(ws, total_row, 13, "", fill=FILL_TOTAL, border=BORDER)
        _cell(ws, total_row, 14, f"{float(total_consumed):.3f}", font=BOLD, fill=FILL_TOTAL, align=CENTER, border=BORDER)
        _cell(ws, total_row, 15, f"{float(total_refill):.1f}",   font=BOLD, fill=FILL_TOTAL, align=CENTER, border=BORDER)
        for cc in range(16, 19):
            _cell(ws, total_row, cc, "", fill=FILL_TOTAL, border=BORDER)

        # ── Ширина колонок ──
        col_widths = [5, 13, 5, 8, 8, 8, 8, 8, 8, 8, 18, 10, 10, 10, 10, 8, 10, 16]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].height = 30
        ws.freeze_panes = "A5"

    # ── АРКУШ 2: ЗВЕДЕННЯ МІСЯЦЯ ───────────────────────────────────────────

    async def _build_monthly_summary(
        self, wb: Workbook, gen_name: str, year: int, month: int,
        shifts: list[Shift], refills: list[FuelRefill],
        deliveries: list[FuelDelivery], fuel_price: float
    ) -> None:
        ws = wb.create_sheet("Зведення місяця")
        month_str  = f"{UA_MONTHS[month]} {year}"
        now_str    = datetime.now(tz=KYIV).strftime("%d.%m.%Y %H:%M")
        days_count = calendar.monthrange(year, month)[1]
        weekends   = sum(1 for d in range(1, days_count + 1) if date(year, month, d).weekday() >= 5)

        # ── Зведені показники ──
        total_hours    = sum((sh.motohours_accumulated or Decimal("0")) for sh in shifts)
        total_consumed = sum((sh.fuel_consumed_liters  or Decimal("0")) for sh in shifts)
        total_refill   = sum((rf.liters for rf in refills), Decimal("0"))
        total_delivered = sum((d.liters for d in deliveries), Decimal("0"))

        # Робочих днів = дні з хоча б однією завершеною зміною
        worked_dates = set()
        for sh in shifts:
            if sh.stopped_at:
                worked_dates.add(sh.started_at.astimezone(KYIV).date())
        worked_days = len(worked_dates)

        avg_lph = (total_consumed / total_hours) if total_hours and total_hours > 0 else Decimal("0")
        fuel_cost = total_consumed * Decimal(str(fuel_price))

        # ── Шапка ──
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Зведення місяця — «{gen_name}»  —  {month_str}"
        ws["A1"].font = HDR1
        ws["A1"].alignment = CENTER

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Сформовано: {now_str}"
        ws["A2"].font = BODY
        ws["A2"].alignment = LEFT

        # ── KPI блок ──
        kpi_row = 4
        _merge(ws, kpi_row, 1, kpi_row, 8, "Ключові показники", font=Font(bold=True, size=12), fill=FILL_HEADER, align=CENTER)
        for cc in range(1, 9):
            ws.cell(kpi_row, cc).font = WHITE
            ws.cell(kpi_row, cc).border = BORDER

        kpis = [
            ("Мотогодини за місяць",    f"{float(total_hours):.2f} год"),
            ("Робочих днів",            f"{worked_days} / {days_count}"),
            ("Вихідних у місяці",       str(weekends)),
            ("Загальна витрата палива", f"{float(total_consumed):.1f} л"),
            ("Загальне поповнення",     f"{float(total_refill):.1f} л"),
            ("Постачань за місяць",     f"{float(total_delivered):.1f} л"),
            ("Середня витрата",         f"{float(avg_lph):.2f} л/год"),
            ("Ціна палива",             f"{fuel_price:.2f} грн/л"),
            ("Вартість палива",         f"{int(fuel_cost)} грн"),
        ]
        for i, (label, val) in enumerate(kpis):
            r = kpi_row + 1 + i
            _merge(ws, r, 1, r, 3, label, font=BODY, fill=FILL_KPI_LBL, align=LEFT)
            for cc in range(1, 4):
                ws.cell(r, cc).border = BORDER
            _merge(ws, r, 4, r, 5, val, font=BOLD, fill=None, align=LEFT)
            for cc in range(4, 6):
                ws.cell(r, cc).border = BORDER
            for cc in range(6, 9):
                _cell(ws, r, cc, "", border=BORDER)

        # ── Щоденна таблиця ──
        tbl_start = kpi_row + len(kpis) + 3
        _merge(ws, tbl_start, 1, tbl_start, 8, "Щоденна зведена таблиця",
               font=Font(bold=True, size=11), fill=FILL_HEADER, align=CENTER)
        for cc in range(1, 9):
            ws.cell(tbl_start, cc).font = WHITE
            ws.cell(tbl_start, cc).border = BORDER

        hdr_row = tbl_start + 1
        daily_cols = ["Дата", "День", "Год.", "Пал.⬆, л", "Пал.⬇, л", "Витрата, л", "Заправка, л", "Примітки"]
        for col, label in enumerate(daily_cols, start=1):
            _cell(ws, hdr_row, col, label, font=WHITE, fill=FILL_SUBHDR, align=CENTER, border=BORDER)

        # Індекс зміни/заправок по даті
        shift_by_date: dict = {}
        for sh in shifts:
            d = sh.started_at.astimezone(KYIV).date()
            shift_by_date.setdefault(d, []).append(sh)

        refill_by_date: dict = {}
        for rf in refills:
            d = rf.refilled_at.astimezone(KYIV).date()
            refill_by_date.setdefault(d, []).append(rf)

        for day_num in range(1, days_count + 1):
            d    = date(year, month, day_num)
            row  = hdr_row + day_num
            wd   = d.weekday()
            fill = FILL_WEEKEND if wd >= 5 else (FILL_ALT if day_num % 2 == 0 else None)

            day_shifts  = shift_by_date.get(d, [])
            day_refills = refill_by_date.get(d, [])

            day_hours    = sum((sh.motohours_accumulated or Decimal("0")) for sh in day_shifts)
            day_consumed = sum((sh.fuel_consumed_liters  or Decimal("0")) for sh in day_shifts)
            day_refill   = sum((rf.liters for rf in day_refills), Decimal("0"))

            row_data = [
                d.strftime("%d.%m.%Y"),
                UA_DAYS[wd],
                f"{float(day_hours):.2f}" if day_hours else "",
                "", "",  # Пал.⬆ / Пал.⬇ — немає в моделях
                f"{float(day_consumed):.3f}" if day_consumed else "",
                f"{float(day_refill):.1f}" if day_refill else "",
                "—",
            ]
            for col, val in enumerate(row_data, start=1):
                _cell(ws, row, col, val, font=BODY, fill=fill, align=CENTER, border=BORDER)

        # Підсумок
        sum_row = hdr_row + days_count + 1
        _merge(ws, sum_row, 1, sum_row, 2, "ПІДСУМОК", font=BOLD, fill=FILL_TOTAL, align=CENTER)
        _cell(ws, sum_row, 3, f"{float(total_hours):.2f}",    font=BOLD, fill=FILL_TOTAL, align=CENTER, border=BORDER)
        _cell(ws, sum_row, 4, "",                             fill=FILL_TOTAL, border=BORDER)
        _cell(ws, sum_row, 5, "",                             fill=FILL_TOTAL, border=BORDER)
        _cell(ws, sum_row, 6, f"{float(total_consumed):.3f}", font=BOLD, fill=FILL_TOTAL, align=CENTER, border=BORDER)
        _cell(ws, sum_row, 7, f"{float(total_refill):.1f}",   font=BOLD, fill=FILL_TOTAL, align=CENTER, border=BORDER)
        _cell(ws, sum_row, 8, "",                             fill=FILL_TOTAL, border=BORDER)
        for cc in range(1, 9):
            ws.cell(sum_row, cc).border = BORDER

        col_widths = [14, 6, 8, 10, 10, 12, 12, 20]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{hdr_row + 1}"

    # ── АРКУШ 3: ТЕХНІЧНЕ ОБСЛУГОВУВАННЯ ───────────────────────────────────

    async def _build_maintenance_sheet(
        self, wb: Workbook, gen_name: str,
        maintenance: list[tuple]
    ) -> None:
        ws = wb.create_sheet("Технічне обслуговування")

        ws.merge_cells("A1:E1")
        ws["A1"] = f"Технічне обслуговування — «{gen_name}»"
        ws["A1"].font = HDR1
        ws["A1"].alignment = CENTER

        # ── Заголовки таблиці ──
        headers = ["Дата", "Тип ТО", "Мотогодини", "Виконав", "Примітки"]
        for col, label in enumerate(headers, start=1):
            _cell(ws, 3, col, label, font=WHITE, fill=FILL_HEADER, align=CENTER, border=BORDER)

        MAINT_TYPE_UA = {
            "SCHEDULED": "Планове ТО",
            "OIL_CHANGE": "Заміна мастила",
            "SPARK_PLUG": "Заміна свічок",
            "FILTER":     "Заміна фільтра",
            "OTHER":      "Інше",
        }

        for row_idx, row_data in enumerate(maintenance, start=4):
            if isinstance(row_data, tuple):
                m, user_name = row_data[0], row_data[1]
            else:
                m = row_data
                user_name = "—"

            fill = FILL_ALT if row_idx % 2 == 0 else None
            maint_type = MAINT_TYPE_UA.get(getattr(m, "maintenance_type", ""), "Технічне ТО")

            row_vals = [
                m.performed_at.astimezone(KYIV).strftime("%d.%m.%Y %H:%M") if m.performed_at else "—",
                maint_type,
                f"{float(m.motohours_at_service):.1f} год" if m.motohours_at_service else "—",
                user_name or "—",
                m.notes or "",
            ]
            for col, val in enumerate(row_vals, start=1):
                _cell(ws, row_idx, col, val, font=BODY, fill=fill, align=LEFT if col >= 3 else CENTER, border=BORDER)

        if not maintenance:
            _merge(ws, 4, 1, 4, 5, "Записів не знайдено", font=BODY, align=CENTER)

        col_widths = [20, 20, 16, 24, 30]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A4"
